from django.core.management.base import BaseCommand, CommandError

from market.models import Category, Product

from project.settings import DATA_DIR  

from openpyxl import load_workbook


class Command(BaseCommand):


    def handle(self, *args, **options):

        print('Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print("start blabla %s" % DATA_DIR)

        wb = load_workbook(DATA_DIR+'/price.xlsx')
        worksheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

        cat = None

        for count in range (1, worksheet.max_row+1):
            item = worksheet.cell(row = count, column = 3) .value
            id = worksheet.cell(row = count, column = 2) .value
            if id == None:
                print('New Category')
                cat = Category()
                cat.name = item
                cat.save()
            else:
                print('New Product')
                if cat:
                    p = Product()
                    p.name = item
                    p.category = cat
                    p.save()

                
 